#For any quries contact : kumarakhil224@gmail.com
#This program requires the .log file obtained from gaussian containing frequencies and intensities. 
#Generates the excel file matching each frequency with its corresponding intensity.

from openpyxl import Workbook

input_file = input("Enter the input file name\n")
output_file = input("Enter the output file name\n")
out_file= output_file + ".xlsx"

freq=[]
inten=[]
with open(input_file,'r') as f:
 for lines in f:
  if "Frequencies" in lines:
   parts=lines.strip().split()
   freq.extend(float(x) for x in parts[2:])
  if "IR Inten" in lines:
   parts=lines.strip().split()
   inten.extend(float(x) for x in parts[3:])

print(freq)
print(inten)

wb = Workbook()
ws = wb.active

ws.cell(row=1,column=1,value="Frequencies")
ws.cell(row=1,column=2,value="Intensities")

for i,value in enumerate(freq,start=2):
 ws.cell(row=i,column=1,value=value)

for i,value in enumerate(inten,start=2):
 ws.cell(row=i,column=2,value=value)

wb.save(out_file)
